import os
import zipfile
import pandas as pd
import re
from openpyxl import load_workbook
import subprocess
import sys

def open_result_file(filepath):
    """Функция для открытия файла с результатами"""
    if sys.platform == "win32":
        os.startfile(filepath)
    elif sys.platform == "darwin":
        subprocess.run(["open", filepath])
    else:
        subprocess.run(["xdg-open", filepath])


def process_zip_files(folder_path, excel_template_path, output_excel_path, org_code='1858'):
    wb = load_workbook(excel_template_path)
    ws = wb.active


    code_to_row = {}
    for row in range(1, ws.max_row + 1):
        code = ws.cell(row=row, column=1).value
        if code is not None:
            code_to_row[str(code).zfill(4)] = row

    pattern = re.compile(r'.*?(\d{4})[^\d]*(\d{4})')

    # Обрабатываем все ZIP-файлы
    for filename in os.listdir(folder_path):
        if filename.endswith('.zip'):
            # Ищем коды в имени файла
            match = pattern.search(filename)
            if not match or len(match.groups()) != 2:
                print(f"Не удалось извлечь коды из файла: {filename}")
                continue

            code1, code2 = match.groups()
            is_org_first = (code1 == org_code)
            counterpart_code = code2 if is_org_first else code1


            if counterpart_code not in code_to_row:
                print(f"Код контрагента {counterpart_code} не найден в Excel файле")
                continue

            row = code_to_row[counterpart_code]
            zip_path = os.path.join(folder_path, filename)

            try:
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    # Обрабатываем в зависимости от того, чей код первый
                    if not is_org_first:
                        # Случай 1: код контрагента первый
                        try:
                            with zip_ref.open('Chapter7 A-B.csv') as ab_file:
                                ab_df = pd.read_csv(ab_file, header=None, skiprows=1, encoding='cp1251', sep=';', decimal=',')
                                # Ищем строку с "Общий итог"
                                total_row = ab_df[ab_df[0].str.contains('Общий итог', na=False)]
                                if not total_row.empty:
                                    # Берем данные из столбцов 1-7 (B-H)
                                    row_data = total_row.iloc[0, 1:8].tolist()
                                    for col_idx, value in enumerate(row_data, start=3):  # Столбцы C-I
                                        if isinstance(value, str):
                                            try:
                                                # Удаляем пробелы между разрядами и заменяем запятую на точку, преобразуем в число
                                                value = float(value.replace(' ', '').replace(',', '.'))
                                            except ValueError:
                                                pass
                                        ws.cell(row=row, column=col_idx, value=value)
                        except Exception as e:
                            print(f"Ошибка при обработке Chapter7 A-B.csv в {filename}: {str(e)}")

                        try:
                            with zip_ref.open('Chapter7 B-A.csv') as ba_file:
                                ba_df = pd.read_csv(ba_file, header=None,skiprows=1, encoding='cp1251', sep=';', decimal=',')
                                # Ищем строку с "Общий итог"
                                total_row = ba_df[ba_df[0].str.contains('Общий итог', na=False)]
                                if not total_row.empty:
                                    # Берем данные из столбцов 1-7 (B-H)
                                    row_data = total_row.iloc[0, 1:8].tolist()
                                    for col_idx, value in enumerate(row_data, start=11):  # Столбцы K-Q
                                        if isinstance(value, str):
                                            try:
                                                # Удаляем пробелы между разрядами и заменяем запятую на точку, преобразуем в число
                                                value = float(value.replace(' ', '').replace(',', '.'))
                                            except ValueError:
                                                pass
                                        ws.cell(row=row, column=col_idx, value=value)
                        except Exception as e:
                            print(f"Ошибка при обработке Chapter7 B-A.csv в {filename}: {str(e)}")
                    else:
                        # Случай 2: код организации первый
                        try:
                            with zip_ref.open('Chapter7 B-A.csv') as ba_file:
                                ba_df = pd.read_csv(ba_file, header=None, skiprows=1, encoding='cp1251', sep=';', decimal=',')
                                # Ищем строку с "Общий итог"
                                total_row = ba_df[ba_df[0].str.contains('Общий итог', na=False)]
                                if not total_row.empty:
                                    # Берем данные из столбцов 1-7 (B-H)
                                    row_data = total_row.iloc[0, 1:8].tolist()
                                    for col_idx, value in enumerate(row_data, start=3):
                                        # Столбцы C-I
                                        if isinstance(value, str):
                                            try:
                                                # Удаляем пробелы между разрядами и заменяем запятую на точку, преобразуем в число
                                                value = float(value.replace(' ', '').replace(',', '.'))
                                            except ValueError:
                                                pass
                                        ws.cell(row=row, column=col_idx, value=value)
                        except Exception as e:
                            print(f"Ошибка при обработке Chapter7 B-A.csv в {filename}: {str(e)}")

                        try:
                            with zip_ref.open('Chapter7 A-B.csv') as ab_file:
                                ab_df = pd.read_csv(ab_file, header=None, skiprows=1, encoding='cp1251', sep=';', decimal=',')
                                # Ищем строку с "Общий итог"
                                total_row = ab_df[ab_df[0].str.contains('Общий итог', na=False)]
                                if not total_row.empty:
                                    # Берем данные из столбцов 1-7 (B-H)
                                    row_data = total_row.iloc[0, 1:8].tolist()
                                    for col_idx, value in enumerate(row_data, start=11):  # Столбцы K-Q
                                        if isinstance(value, str):
                                            try:
                                                # Удаляем пробелы между разрядами и заменяем запятую на точку, преобразуем в число
                                                value = float(value.replace(' ', '').replace(',', '.'))
                                            except ValueError:
                                                pass
                                        ws.cell(row=row, column=col_idx, value=value)
                        except Exception as e:
                            print(f"Ошибка при обработке Chapter7 A-B.csv в {filename}: {str(e)}")
            except Exception as e:
                print(f"Ошибка при обработке файла {filename}: {str(e)}")

    # Сохраняем результат
    try:
        wb.save(output_excel_path)
        print(f"Файл успешно сохранён по пути: {os.path.abspath(output_excel_path)}")
        return True
    except Exception as e:
        print(f"Ошибка сохранения: {str(e)}")
        return False


def main():

    folder_path = input("Введите путь к папке с ZIP-архивами: ").strip('"')

    # Определяем пути к файлам относительно исполняемого файла
    if getattr(sys, 'frozen', False):
        base_dir = os.path.dirname(sys.executable)
    else:
        base_dir = os.path.dirname(os.path.abspath(__file__))

    excel_template_path = os.path.join(base_dir, 'Шаблон.xlsx')
    output_excel_path = 'Результат.xlsx'

    if not os.path.exists(excel_template_path):
        print("\nОШИБКА: Не найден файл шаблона 'Шаблон.xlsx'")
        print("Убедитесь, что он находится в той же папке, что и программа")
        input("Нажмите Enter для выхода...")
        return

    print("\nОбработка данных...")

    # Вызываем функцию обработки (это было пропущено в вашем коде)
    success = process_zip_files(folder_path, excel_template_path, output_excel_path)

    if success:
        print("\nГотово! Открываю файл с результатами...")
        try:
            open_result_file(output_excel_path)
        except Exception as e:
            print(f"Не удалось открыть файл автоматически: {str(e)}")
            print(f"Результаты сохранены в: {output_excel_path}")
    else:
        print("\nНе удалось сохранить результаты. Проверьте сообщения об ошибках выше.")

    input("Нажмите Enter для выхода...")


if __name__ == "__main__":
    main()

